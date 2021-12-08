from os import walk
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import glob
import datetime

# develop by Min CAI  7/12/2021


class Transaction:
    def __init__(self,team,customer,inorout,approve_state,dayOftransaction,amount,outbythisday):
        self.team = team
        self.customer = customer
        self.inorout = inorout
        self.approve_state = approve_state
        self.dayOftransaction = dayOftransaction
        self.amount = amount
        self.outbythisday = outbythisday

class Customer:
    finalResultAmount = 0

    def __init__(self, name, transaction):
        self.name = name
        self.transaction = transaction


class reportList:
    def __init__(self,companyName, resultList,finalPerDayAmount):
        self.companyName = companyName
        self.resultList = resultList
        self.finalPerDayAmount = finalPerDayAmount



class AmountPerDay:
    def __init__(self,companyName, dayrate, thisday):
        self.dayrate = dayrate
        self.thisday = thisday
        self.companyName = companyName


def readData(row):
    team = sheet['C' + str(row)].value
    customer = sheet['E' + str(row)].value
    inorout = sheet['H' + str(row)].value
    approve_state = sheet['I' + str(row)].value
    dayOftransaction = sheet['L' + str(row)].value
    amount = sheet['Y' + str(row)].value
    outbythisday = sheet['Z' + str(row)].value
    return Transaction(team,customer,inorout,approve_state,dayOftransaction,amount-outbythisday,outbythisday)


xlsxfiles = []
for file in glob.glob("*.xlsx"):
    xlsxfiles.append(file)
print(xlsxfiles)

# reading the day and time
wb = openpyxl.load_workbook("11.30.xlsx", data_only=True)
sheet = wb.active
target_day = sheet['CM1'].value
first_day = sheet['CM2'].value
days = sheet['CM3'].value
print("Target Day: ", target_day)
print("Frist Day of the year: ",first_day)
print("Days: ", days)

row = 2
transaction = []
customers = []
finalResultListPerCustomer = []


while row < sheet.max_row:
        if sheet['E' + str(row)].value == sheet['E' + str(row+1)].value:
            transaction.append(readData(row))
        elif sheet['E' + str(row)].value != sheet['E' + str(row+1)].value:
            transaction.append(readData(row))
            transaction.sort(key=lambda x: x.dayOftransaction)
            company_name = transaction[0].customer  
            customers.append(Customer(company_name,transaction))
            print ("reading date in Customer: " ,company_name)
            transaction = []
            #break 3 customer for developing here
            #if len(customers) == 3:
                #break

        row += 1
        


dayrecord = []
daysForCal = 0

if days == None:
    #calculate the day
    daysForCal = (target_day - first_day).days
    print("Does not provided any days")
else:
    daysForCal = days
    print("Provided recording days")



lengthOfDaysForTracking = 0

print ("There are ", len(customers), "customers in this document")
processpercentage = 0

for customer in customers:
    #calculate the days between first transaction and target day
    tempresultList = []
    sum = 0
    perdayRateAmount = 0
    currentdate = customer.transaction[0].dayOftransaction
    lengthOfDaysForTracking = (target_day - currentdate).days  
    for dayIndex in range(0,lengthOfDaysForTracking+1):
        for trans in customer.transaction:
            if currentdate == trans.dayOftransaction:
                perdayRateAmount = perdayRateAmount + trans.amount
                if perdayRateAmount < 0:
                    perdayRateAmount = 0
                temp_dayrate = perdayRateAmount / daysForCal
        tempresultList.append(AmountPerDay(customer.name,temp_dayrate,currentdate))
        sum = sum + temp_dayrate
        currentdate += datetime.timedelta(days=1)

    finalResultListPerCustomer.append(reportList(customer.name, tempresultList,sum))
    print("Dealing with cusotmer: ", customer.name, " , remaining: ", len(customers) - processpercentage)
    processpercentage += 1
    tempresultList = []
    sum = 0
    perdayRateAmount = 0
    


for result in finalResultListPerCustomer:
    print("companyName: ", result.companyName, end= " ")
    #for perdayList in result.resultList:
        #print("companyName: ", result.companyName, end= " ")
        #print("Date: ", perdayList.thisday,end=" ")
        #print("Per Day rate: ", perdayList.dayrate)

    print("Sum :", result.finalPerDayAmount)

# create a result excel sheet
nwb = Workbook()
ws = nwb.active
ws.title = "Result report"
print_row = 1
ws['A1'] = "公司编号"
ws['B1'] = "公司名称"
ws['C1'] = "日均总和"
print_row += 1
# print result
for result in finalResultListPerCustomer:
    ws['A' + str(print_row)] = print_row - 1
    ws['B' + str(print_row)] = result.companyName
    ws['C' + str(print_row)] = result.finalPerDayAmount
    print_row += 1

nwb.save('Report.xlsx')




# print detail report
nwb2 = Workbook()
ws2 = nwb2.active
ws2.title = "Result report detail"
companyIndex = 1
print_row = 1
print_column = 1

for result in finalResultListPerCustomer:
    ws2[get_column_letter(print_column) + str(print_row)] = "公司编号"
    ws2[get_column_letter(print_column+1) + str(print_row)] = companyIndex
    ws2[get_column_letter(print_column+2) + str(print_row)] = "公司名称"
    ws2[get_column_letter(print_column+3) + str(print_row)] = result.companyName
    ws2[get_column_letter(print_column+4) + str(print_row)] = "日均总和"
    ws2[get_column_letter(print_column+5) + str(print_row)] = result.finalPerDayAmount  
    companyIndex += 1   
    print_row += 1

    ws2[get_column_letter(print_column) + str(print_row)] = "日期"
    ws2[get_column_letter(print_column+1) + str(print_row)] = "日均"
    print_row += 1
    for perdayList in result.resultList:
        #print("Date: ", perdayList.thisday,end=" ")
        #print("Per Day rate: ", perdayList.dayrate)
        char = get_column_letter(print_column)
        char2 = get_column_letter(print_column+1)
        ws2[char + str(print_row)] = perdayList.thisday.strftime('%Y-%m-%d')
        ws2[char2 + str(print_row)] = perdayList.dayrate
        print_row += 1
    print_column += 7
    print_row = 1

nwb2.save('report detail.xlsx')



     


    


